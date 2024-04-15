#import docx as docx
import openpyxl
import datetime
import sys
import os
import argparse

CELL=[]
CELL.append(("申請年",15,31))
CELL.append(("申請月",15,36))
CELL.append(("申請日",15,42))
CELL.append(("職",23,14))
CELL.append(("氏名",24,14))
CELL.append(("用務内容",33,3))
CELL.append(("全ての用務先",37,3))

CELL.append(("旅行開始年",39,5))
CELL.append(("旅行開始月",39,7))
CELL.append(("旅行開始日",39,9))
CELL.append(("旅行終了年",39,14))
CELL.append(("旅行終了月",39,17))
CELL.append(("旅行終了日",39,23))
CELL.append(("旅行日数",39,32))
CELL.append(("経費",64,8))
CELL.append(("備考",65,3))
CELL.append(("安全保障",67,8))

TABLEROW=[43,46,49,52,55,58,61]
TABLECELL=[("YYYY/MM/DD",2),("出発地",5),("到着地",9),("用務先",13),("宿泊及び滞在地",28),("宿泊数",41)]

SHEETNAME="別紙様式１"


def make_xlsx(basexlsx,outxlsx,formdata):
    if basexlsx==outxlsx:
        return
    wb = openpyxl.load_workbook(basexlsx)
    ws = wb[SHEETNAME]

    for (k,i,j) in CELL:
        if k in formdata:
            ws.cell(row=i,column=j,value=formdata[k])
            
    for (ni,i) in zip(formdata["日程"],TABLEROW):
        for k,j in TABLECELL:
            if k in ni:
                ws.cell(row=i,column=j,value=ni[k])

    wb.save(outxlsx)

def get_wareki_int(yyyy):
    return yyyy-2018

def get_wareki_str(yyyy,mm,dd):
    return "令和{0}年{1}月{2}日".format(get_wareki_int(yyyy),mm,dd)

def get_wareki_short_str(yyyy,mm,dd):
    return "R{0}.{1}.{2}".format(get_wareki_int(yyyy),mm,dd)

def read_formdata_from_file(file):
    with open(file) as f:
        data={"":"","日程":[]}
        currenttag=""
        keys=["申請年月日","職","氏名","用務内容","経費","備考","安全保障","mailto_name","mailto_address","filenameprefix"]
        tabelKeys=["出発地","到着地","用務先","宿泊及び滞在地","宿泊数"]
        
        for l in f:
            li=l.strip()
            if li.startswith("#"):
                continue
            if li.startswith("%"):
                continue
            if not li.startswith(":"):
                if currenttag=="日程":
                    data[currenttag][-1][subcurrenttag]=data[currenttag][-1][subcurrenttag]+li
                else:
                    data[currenttag]=data[currenttag]+li
            for k in keys:
                if li.startswith(":"+k+":"):
                    currenttag=k
                    data[currenttag]=li[len(k)+2:]
            if li.startswith(":年月日:"):
                currenttag="日程"
                data[currenttag].append({})
                subcurrenttag="年月日"
                data[currenttag][-1][subcurrenttag]=li[5:]
            for k in tabelKeys:
                if li.startswith(":"+k+":"):
                    currenttag="日程"
                    subcurrenttag=k
                    data[currenttag][-1][subcurrenttag]=li[len(k)+2:]

        formdata={}
        if "申請年月日" in data:
            d=[ int(di) for di in data["申請年月日"].split("/")]
            formdata["申請年月日"]=get_wareki_str(d[0],d[1],d[2])
            formdata["申請年"]=get_wareki_int(d[0])
            formdata["申請月"]=d[1]
            formdata["申請日"]=d[2]
        else:
            dt_now = datetime.datetime.now()
            formdata["申請年月日"]=get_wareki_str(dt_now.year,dt_now.month,dt_now.day)
            formdata["申請年"]=get_wareki_int(dt_now.year)
            formdata["申請月"]=dt_now.month
            formdata["申請日"]=dt_now.day

        if data["日程"] != []:
            date=None
            firstdate=None
            formdata["日程"]=[]
            for i,ni in enumerate(data["日程"]):
                nni={}
                for k in ni.keys():
                    nni[k]=ni[k]
                if ni["年月日"] != "":
                    d=[ int(di) for di in ni["年月日"].split("/")]
                    nni["年月日"]=get_wareki_short_str(d[0],d[1],d[2])
                    nni["YYYY/MM/DD"]="{0:4d}/{1:02d}/{2:02d}".format(d[0],d[1],d[2])
                    dt=datetime.datetime(year=d[0], month=d[1], day=d[2])
                    nni["datetime"]=dt
                    if date==None:
                        date=dt
                        firstdate=dt
                    elif date != dt:
                        formdata["日程"][-1]["宿泊数"]="{0}".format((dt-date).days)
                        date=dt
                        
                formdata["日程"].append(nni)
            d0=get_wareki_str(firstdate.year,firstdate.month,firstdate.day)
            d1=get_wareki_str(date.year,date.month,date.day)
            d2=1+(date-firstdate).days
            formdata["旅行開始年"]=get_wareki_int(firstdate.year)
            formdata["旅行開始月"]=firstdate.month
            formdata["旅行開始日"]=firstdate.day
            formdata["旅行終了年"]=get_wareki_int(date.year)
            formdata["旅行終了月"]=date.month
            formdata["旅行終了日"]=date.day
            formdata["旅行日数"]=d2
            
            formdata["旅行期間"]="自 {0} 〜 至 {1} ({2}日間)".format(d0,d1,d2)
            formdata["day0"]=d0
            if firstdate.year==date.year:
                if firstdate.month==date.month:
                    if firstdate.day==date.day:
                        lastdate=""
                    else:
                        lastdate="-{2:02d}".format(date.year,date.month,date.day)
                        formdata["day1"]="{2:2d}日".format(date.year,date.month,date.day)
                else:
                    lastdate="-{1:02d}/{2:02d}".format(date.year,date.month,date.day)
                    formdata["day1"]="{1:2d}月{2:2d}日".format(date.year,date.month,date.day)
            else:
                lastdate="-{0:4d}/{1:02d}/{2:02d}".format(date.year,date.month,date.day)
                formdata["day1"]="{0:4d}年{1:2d}月{2:2d}日".format(date.year,date.month,date.day)
                    
            formdata["YYYY/MM/DD-"]="{0:4d}/{1:02d}/{2:02d}".format(firstdate.year,firstdate.month,firstdate.day)+lastdate
            formdata["YYYYMMDD"]="{0:4d}{1:02d}{2:02d}".format(firstdate.year,firstdate.month,firstdate.day)
            formdata["全ての用務先"]=""
            for ni in data["日程"]:
                if "用務先" in ni:
                    if formdata["全ての用務先"] != "":
                        formdata["全ての用務先"]=formdata["全ての用務先"]+", "
                    formdata["全ての用務先"]=formdata["全ての用務先"]+ni["用務先"]

        keys_justcopy=["用務内容","mailto_name","mailto_address","filenameprefix","安全保障","経費","職", "氏名","備考"]
        for k in keys_justcopy:
            if k in data:
                formdata[k]=data[k].replace("\\\\","\n")
        
        return formdata

    return None


def mailtxt(formdata):
    r=""
    if "mailto_address" in formdata:
        r="To: "+formdata["mailto_address"]+"\n"
    r=r+"Subject: 他機関経費での出張"
    if "YYYY/MM/DD-" in formdata:
        r=r+" ("+formdata["YYYY/MM/DD-"]+")"
    r=r+"\n\n"
    if "mailto_name" in formdata:
        r=r+formdata["mailto_name"]+",\n\n"
    if "研修内容" in formdata:
        r=r+"以下の用務のため, "
    if "day0" in formdata:
        if "day1" in formdata:
            r=r+"\n"
            r=r+" "+formdata["day0"]+"から"+formdata["day1"]+"\n"
            r=r+"の日程で, "
        else:
            r=r+"\n"
            r=r+" "+formdata["day0"]+"\n"
            r=r+"に, "
    if "主たる用務先" in formdata:
        r=r+"\n"
        r=r+" "+formdata["主たる用務先"]+"\n"
        r=r+"へ, "
    if "用務内容" in formdata:
        r=r+"他機関の経費で出張いたします:\n"
        r=r+" "+formdata["用務内容"]
        r=r+"\n\n"
    else:
        r=r+"他機関の経費で出張いたします.\n"

    if "経費" in formdata:
        r=r+"旅費等は以下から支弁される予定です:\n"
        r=r+" "+formdata["経費"]+"\n\n"
    r=r+"ファイルを添付いたしますので, お手続きをお願いいたします.\n\n" 
    if "氏名" in formdata:
        r=r+"   "+formdata["氏名"]+"\n" 
    
    return r

def test():
    parser = argparse.ArgumentParser(description='研修用書類生成') 
    parser.add_argument('inputfile', help='研修の内容に関する入力ファイル')
    parser.add_argument('-ox', '--output-xlsx', help='出力するxlsxのファイル名.  デフォルトは出発日を使い, filenameprefix-YYYYMMDD.xlsx')
    parser.add_argument('-om', '--output-mail', help='出力するmailのファイル名. デフォルトは出発日を使い, filenameprefix-YYYYMMDD.txt')
    parser.add_argument('-O', '--output-base', help='出力ファイルを, OUTPUT_BASE.xlsx と OUTPUT_BASE.txt に指定.')
    parser.add_argument('-b', '--basefile', help='元となるxlsx. デフォルトは, minimum.xlsx')   
    args = parser.parse_args()

    inputfile=args.inputfile
    fd=read_formdata_from_file(inputfile)

    if args.basefile:
        basexlsx = args.basefile
    else:
        basexlsx = os.path.join(os.path.dirname(__file__),"minimum.xlsx")
        
    if args.output_xlsx:
        outputxlsx=args.output_xlsx
    elif args.output_base:
            outputxlsx=args.output_base+".xlsx"
    elif "YYYYMMDD" in fd:
        if "filenameprefix" in fd:
            outputxlsx=fd["filenameprefix"]+"-"+fd["YYYYMMDD"]+".xlsx"
        else:
            outputxlsx=fd["YYYYMMDD"]+".xlsx"
    else:
        outputxlsx="test.xlsx"

    if basexlsx==outputxlsx:
        return

    make_xlsx(basexlsx,outputxlsx,fd)
        
    if args.output_mail:
        outputmail=args.output_mail
    elif args.output_base:
            outputmail=args.output_base+".txt"
    elif "YYYYMMDD" in fd:
        if "filenameprefix" in fd:
            outputmail=fd["filenameprefix"]+"-"+fd["YYYYMMDD"]+".txt"
        else:
            outputmail=fd["YYYYMMDD"]+".txt"
    else:
        outputmail="test.txt"

    if inputfile == outputmail:
        return            
    with open(outputmail,"w") as f:
        f.write(mailtxt(fd))

    if "申請年" in fd and "研修開始年" in fd:
        dt0=datetime.datetime(year=fd["申請年"], month=fd["申請月"], day=fd["申請日"])
        dt1=datetime.datetime(year=fd["研修開始年"], month=fd["研修開始月"], day=fd["研修開始日"])
        if dt0>dt1:
            print("WARNING: Please check the date.")



if __name__ == "__main__":
    test()


